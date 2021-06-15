#미국 3대 증권사 전종목 가져오기

import FinanceDataReader as fdr     # 모듈 설치 : pip install -U finance-datareader / -U는 나중에 옵구레이드할 때 사용
import openpyxl

df_NASDAQ = fdr.StockListing('NASDAQ')    # 나스닥 전종목 구하기
# print(df_NASDAQ)
df_NASDAQ.to_excel("nasdaq.xlsx")
df_NYSE = fdr.StockListing('NYSE')    # 뉴욕증권거래소 전종목 구하기
# print(df_NYSE)
df_NYSE.to_excel("nyse.xlsx")
df_AMEX = fdr.StockListing('AMEX')    # 아멕스 전종목 구하기
print(df_AMEX)
df_AMEX.to_excel("amex.xlsx")
# wb = openpyxl.Workbook()        # 임시로 엑셀파일(워크북 타입 값) 생성

# sheet = wb.active       # 활성화된 sheet 호출
# sheet.title = 'NASDAQ'
# sheet1 = wb.create_sheet('NYSE')
# sheet2 = wb.create_sheet('AMEX')
# df_NASDAQ = fdr.StockListing('NASDAQ')
# sheet.append(df_NASDAQ)
# wb.save("usa_stock_name.xlsx")  # save 명령으로 실제 파일 저장